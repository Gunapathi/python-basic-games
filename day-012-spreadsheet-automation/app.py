import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
# cell = sheet['a1']
cell = sheet.cell(1, 1)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    correct_value = cell.value * 0.9
    correct_price_cell = sheet.cell(row, 4)
    correct_price_cell.value = correct_value

wb.save('transcations2.xlsx')

# ADDITIONAL - START

# print(sheet.value)
# print(sheet.max_row)

# ADDITIONAL - END